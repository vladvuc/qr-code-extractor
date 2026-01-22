#!/usr/bin/env python3
"""
QR Code Detection and Extraction Script
Processes Excel file with Firebase Storage URLs to detect QR codes and extract alphanumeric codes
"""

import openpyxl
import requests
from PIL import Image
from io import BytesIO
import re
from datetime import datetime
import time
from typing import Dict, Tuple, Optional
import traceback

# Try to import QR code libraries
try:
    from pyzbar.pyzbar import decode as pyzbar_decode
    PYZBAR_AVAILABLE = True
except ImportError:
    PYZBAR_AVAILABLE = False
    print("Warning: pyzbar not available, QR code decoding will be limited")

try:
    import cv2
    import numpy as np
    CV2_AVAILABLE = True
except ImportError:
    CV2_AVAILABLE = False
    print("Warning: cv2 not available, advanced image processing will be limited")

try:
    from pyzxing import BarCodeReader
    PYZXING_AVAILABLE = True
except ImportError:
    PYZXING_AVAILABLE = False
    print("Warning: pyzxing not available")


class QRCodeProcessor:
    def __init__(self, input_file: str, output_file: str):
        self.input_file = input_file
        self.output_file = output_file
        self.results = {
            'total': 0,
            'successful': 0,
            'qr_found': 0,
            'codes_extracted': 0,
            'errors': 0,
            'error_details': []
        }
        self.processing_log = []

    def fetch_image(self, url: str, timeout: int = 30) -> Optional[Image.Image]:
        """Fetch image from URL with error handling"""
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            response = requests.get(url, timeout=timeout, headers=headers, stream=True)
            response.raise_for_status()

            # Load image
            image = Image.open(BytesIO(response.content))
            return image

        except requests.exceptions.Timeout:
            raise Exception(f"Timeout fetching image (>{timeout}s)")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Network error: {str(e)}")
        except Exception as e:
            raise Exception(f"Image loading error: {str(e)}")

    def decode_qr_code(self, image: Image.Image) -> Tuple[bool, Optional[str], Optional[str]]:
        """
        Detect QR code and extract data
        Returns: (qr_detected, decoded_data, extracted_code)
        """
        qr_detected = False
        decoded_data = None
        extracted_code = None

        # Method 1: Try pyzbar
        if PYZBAR_AVAILABLE:
            try:
                decoded_objects = pyzbar_decode(image)
                if decoded_objects:
                    qr_detected = True
                    for obj in decoded_objects:
                        if obj.type == 'QRCODE':
                            decoded_data = obj.data.decode('utf-8', errors='ignore')
                            # Try to extract alphanumeric code from decoded data
                            match = re.search(r'\b[A-Z0-9]{10}\b', decoded_data)
                            if match:
                                extracted_code = match.group(0)
                            break
            except Exception as e:
                print(f"pyzbar decode error: {e}")

        # Method 2: Try OpenCV if pyzbar failed or not available
        if not qr_detected and CV2_AVAILABLE:
            try:
                # Convert PIL to OpenCV format
                img_array = np.array(image)
                if len(img_array.shape) == 3 and img_array.shape[2] == 3:
                    img_cv = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
                else:
                    img_cv = img_array

                # Try QRCodeDetector
                qr_detector = cv2.QRCodeDetector()
                data, bbox, straight_qrcode = qr_detector.detectAndDecode(img_cv)

                if data:
                    qr_detected = True
                    decoded_data = data
                    # Try to extract alphanumeric code
                    match = re.search(r'\b[A-Z0-9]{10}\b', data)
                    if match:
                        extracted_code = match.group(0)
                elif bbox is not None:
                    # QR code detected but not decoded
                    qr_detected = True

            except Exception as e:
                print(f"OpenCV decode error: {e}")

        # Method 3: Try to extract visible text from image using OCR-like pattern matching
        # This would require pytesseract, but we'll use image analysis instead
        if qr_detected and not extracted_code:
            # Try pattern matching on any decoded data
            if decoded_data:
                # Look for common QR code patterns
                patterns = [
                    r'\b([A-Z0-9]{10})\b',  # 10 character alphanumeric
                    r'code[:\s]*([A-Z0-9]{8,12})',  # Code with label
                    r'id[:\s]*([A-Z0-9]{8,12})',  # ID with label
                ]
                for pattern in patterns:
                    match = re.search(pattern, decoded_data, re.IGNORECASE)
                    if match:
                        extracted_code = match.group(1).upper()
                        break

        return qr_detected, decoded_data, extracted_code

    def process_image(self, url: str, row_num: int) -> Dict:
        """Process a single image"""
        result = {
            'row': row_num,
            'url': url,
            'sticker_detected': False,
            'qr_code': '',
            'error': None,
            'decoded_data': None
        }

        try:
            # Fetch image
            image = self.fetch_image(url)

            if image is None:
                raise Exception("Failed to load image")

            # Detect and decode QR code
            qr_detected, decoded_data, extracted_code = self.decode_qr_code(image)

            result['sticker_detected'] = qr_detected
            result['decoded_data'] = decoded_data

            if extracted_code:
                result['qr_code'] = extracted_code
            elif decoded_data:
                # If we have decoded data but no extracted code, use the decoded data
                result['qr_code'] = decoded_data[:50]  # Limit to 50 chars

            if qr_detected:
                self.results['qr_found'] += 1
            if extracted_code:
                self.results['codes_extracted'] += 1

            self.results['successful'] += 1

        except Exception as e:
            result['error'] = str(e)
            self.results['errors'] += 1
            self.results['error_details'].append({
                'row': row_num,
                'url': url,
                'error': str(e)
            })

        return result

    def process_excel(self):
        """Main processing function"""
        print(f"Loading Excel file: {self.input_file}")
        wb = openpyxl.load_workbook(self.input_file)
        ws = wb.active

        # Find column indices
        headers = [cell.value for cell in ws[1]]
        print(f"Found headers: {headers}")

        # Find URL column (case-insensitive)
        url_col = None
        for idx, header in enumerate(headers, 1):
            if header and 'image' in str(header).lower():
                url_col = idx
                break

        if url_col is None:
            # Try other common names
            for idx, header in enumerate(headers, 1):
                if header and any(name in str(header).lower() for name in ['url', 'link', 'photo']):
                    url_col = idx
                    break

        if url_col is None:
            raise Exception(f"Could not find image URL column. Headers: {headers}")

        # Find or create output columns
        sticker_col = None
        qr_code_col = None

        for idx, header in enumerate(headers, 1):
            if header and 'okret' in str(header).lower() and 'sticker' in str(header).lower():
                sticker_col = idx
            if header and 'qr' in str(header).lower() and 'code' in str(header).lower():
                qr_code_col = idx

        # Create columns if they don't exist
        if sticker_col is None:
            sticker_col = len(headers) + 1
            ws.cell(row=1, column=sticker_col, value='Photo Okret Sticker')
            print(f"Created 'Photo Okret Sticker' column at position {sticker_col}")

        if qr_code_col is None:
            qr_code_col = len(headers) + 2 if sticker_col == len(headers) + 1 else len(headers) + 1
            ws.cell(row=1, column=qr_code_col, value='QR_CODE')
            print(f"Created 'QR_CODE' column at position {qr_code_col}")

        print(f"\nColumn mapping:")
        print(f"  URL column: {url_col} ({headers[url_col-1]})")
        print(f"  Sticker column: {sticker_col}")
        print(f"  QR Code column: {qr_code_col}")

        # Count total rows
        total_rows = ws.max_row - 1  # Exclude header
        self.results['total'] = total_rows

        print(f"\nProcessing {total_rows} images...")
        print("=" * 80)

        start_time = time.time()

        # Process each row
        for row_num in range(2, ws.max_row + 1):
            url_cell = ws.cell(row=row_num, column=url_col)
            url = url_cell.value

            if not url:
                print(f"Row {row_num}: No URL found, skipping")
                continue

            # Progress indicator
            progress = ((row_num - 1) / total_rows) * 100
            print(f"\n[{progress:.1f}%] Processing row {row_num}/{ws.max_row}...")
            print(f"URL: {url[:80]}...")

            # Process the image
            result = self.process_image(url, row_num)
            self.processing_log.append(result)

            # Update Excel cells
            ws.cell(row=row_num, column=sticker_col, value=result['sticker_detected'])
            ws.cell(row=row_num, column=qr_code_col, value=result['qr_code'])

            # Print result
            if result['error']:
                print(f"  ERROR: {result['error']}")
            else:
                print(f"  QR Detected: {result['sticker_detected']}")
                if result['qr_code']:
                    print(f"  Code Extracted: {result['qr_code']}")
                if result['decoded_data'] and result['decoded_data'] != result['qr_code']:
                    print(f"  Full QR Data: {result['decoded_data'][:100]}")

            # Save periodically (every 50 rows)
            if row_num % 50 == 0:
                print(f"\n>>> Saving progress checkpoint at row {row_num}...")
                wb.save(self.output_file)

        # Final save
        print(f"\n\nSaving final results to: {self.output_file}")
        wb.save(self.output_file)

        elapsed_time = time.time() - start_time
        print(f"\nProcessing complete in {elapsed_time:.1f} seconds")

        return wb

    def generate_report(self):
        """Generate detailed processing report"""
        print("\n" + "=" * 80)
        print("PROCESSING REPORT")
        print("=" * 80)
        print(f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"\nInput File: {self.input_file}")
        print(f"Output File: {self.output_file}")

        print(f"\n--- SUMMARY STATISTICS ---")
        print(f"Total Images: {self.results['total']}")
        print(f"Successfully Processed: {self.results['successful']}")
        print(f"QR Codes Detected: {self.results['qr_found']}")
        print(f"Alphanumeric Codes Extracted: {self.results['codes_extracted']}")
        print(f"Errors/Failed: {self.results['errors']}")

        if self.results['total'] > 0:
            success_rate = (self.results['successful'] / self.results['total']) * 100
            qr_rate = (self.results['qr_found'] / self.results['total']) * 100
            extraction_rate = (self.results['codes_extracted'] / self.results['total']) * 100

            print(f"\nSuccess Rate: {success_rate:.1f}%")
            print(f"QR Detection Rate: {qr_rate:.1f}%")
            print(f"Code Extraction Rate: {extraction_rate:.1f}%")

        # Successful extractions
        print(f"\n--- SUCCESSFULLY PROCESSED IMAGES ---")
        successful = [r for r in self.processing_log if not r['error'] and r['qr_code']]
        print(f"Total: {len(successful)}")

        if successful:
            print("\nSample successful extractions (first 10):")
            for i, result in enumerate(successful[:10], 1):
                print(f"{i}. Row {result['row']}: {result['qr_code']}")

        # Error breakdown
        if self.results['error_details']:
            print(f"\n--- ERROR BREAKDOWN ---")
            error_types = {}
            for error in self.results['error_details']:
                error_msg = error['error']
                # Categorize errors
                if 'timeout' in error_msg.lower():
                    category = 'Timeout'
                elif 'network' in error_msg.lower() or 'connection' in error_msg.lower():
                    category = 'Network Error'
                elif '404' in error_msg or 'not found' in error_msg.lower():
                    category = '404 Not Found'
                elif '403' in error_msg or 'forbidden' in error_msg.lower():
                    category = '403 Forbidden'
                else:
                    category = 'Other'

                error_types[category] = error_types.get(category, 0) + 1

            print("\nError categories:")
            for category, count in sorted(error_types.items(), key=lambda x: x[1], reverse=True):
                print(f"  {category}: {count}")

            print(f"\nFailed images (first 10):")
            for i, error in enumerate(self.results['error_details'][:10], 1):
                print(f"{i}. Row {error['row']}: {error['error']}")
                print(f"   URL: {error['url'][:80]}...")

        # Images without QR codes
        no_qr = [r for r in self.processing_log if not r['error'] and not r['sticker_detected']]
        if no_qr:
            print(f"\n--- IMAGES WITHOUT QR CODES ---")
            print(f"Total: {len(no_qr)}")
            print("\nFirst 10 rows without QR codes:")
            for i, result in enumerate(no_qr[:10], 1):
                print(f"{i}. Row {result['row']}")

        print("\n" + "=" * 80)

        # Save detailed report to file
        report_file = self.output_file.replace('.xlsx', '_report.txt')
        with open(report_file, 'w') as f:
            f.write("=" * 80 + "\n")
            f.write("DETAILED PROCESSING REPORT\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write(f"Input File: {self.input_file}\n")
            f.write(f"Output File: {self.output_file}\n\n")

            f.write("--- SUMMARY ---\n")
            f.write(f"Total Images: {self.results['total']}\n")
            f.write(f"Successfully Processed: {self.results['successful']}\n")
            f.write(f"QR Codes Detected: {self.results['qr_found']}\n")
            f.write(f"Alphanumeric Codes Extracted: {self.results['codes_extracted']}\n")
            f.write(f"Errors: {self.results['errors']}\n\n")

            if successful:
                f.write("--- ALL SUCCESSFUL EXTRACTIONS ---\n")
                for result in successful:
                    f.write(f"Row {result['row']}: {result['qr_code']}\n")
                f.write("\n")

            if self.results['error_details']:
                f.write("--- ALL ERRORS ---\n")
                for error in self.results['error_details']:
                    f.write(f"Row {error['row']}: {error['error']}\n")
                    f.write(f"  URL: {error['url']}\n\n")

        print(f"\nDetailed report saved to: {report_file}")


def main():
    input_file = 'EXCEL/Test_Sort_V2.xlsx'
    output_file = 'EXCEL/Test_Sort_V2_Analyzed.xlsx'

    print("=" * 80)
    print("QR CODE DETECTION AND EXTRACTION TOOL")
    print("=" * 80)
    print(f"\nLibraries available:")
    print(f"  pyzbar: {PYZBAR_AVAILABLE}")
    print(f"  OpenCV: {CV2_AVAILABLE}")
    print(f"  pyzxing: {PYZXING_AVAILABLE}")
    print()

    processor = QRCodeProcessor(input_file, output_file)

    try:
        processor.process_excel()
        processor.generate_report()

        print(f"\n✓ Processing complete!")
        print(f"✓ Results saved to: {output_file}")

    except Exception as e:
        print(f"\nFATAL ERROR: {e}")
        traceback.print_exc()
        return 1

    return 0


if __name__ == '__main__':
    exit(main())
